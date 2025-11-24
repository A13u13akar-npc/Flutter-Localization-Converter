import json
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook, load_workbook


def strip_comments(text):
    text = re.sub(r'/\*.*?\*/', '', text, flags=re.DOTALL)
    out = []
    for line in text.splitlines():
        line = re.sub(r'//.*', '', line)
        if line.strip():
            out.append(line)
    return "\n".join(out)


def normalize(s):
    s = s.replace("\\n", "\n")
    s = s.replace("\n", "\\n")
    return s


def extract_dart_strings(path):
    with open(path, "r", encoding="utf-8") as f:
        t = strip_comments(f.read())
    pattern = r'"([^"\$]+)"\.tr'
    out = []
    for m in re.findall(pattern, t):
        out.append(normalize(m))
    return out


def extract_json_keys(path):
    with open(path, "r", encoding="utf-8") as f:
        t = strip_comments(f.read())
    data = json.loads(t)
    return list(data.keys())


def ask_languages_popup(root):
    win = tk.Toplevel(root)
    win.title("Language Codes")
    win.geometry("420x450")
    win.attributes("-topmost", True)

    frame = tk.Frame(win)
    frame.pack(expand=True, fill="both", padx=10, pady=10)

    txt = tk.Text(frame, height=20)
    txt.pack(expand=True, fill="both")

    val = tk.StringVar()

    def done(event=None):
        val.set(txt.get("1.0", tk.END))
        win.destroy()

    tk.Button(win, text="OK", width=12, command=done).pack(pady=10)

    win.bind("<Control-Return>", done)
    win.bind("<Command-Return>", done)

    root.wait_window(win)

    raw = val.get()
    raw = re.sub(r'//.*', '', raw)
    raw = re.sub(r'/\*.*?\*/', '', raw, flags=re.DOTALL)

    codes = []
    codes += re.findall(r"Locale\('([a-zA-Z\-]+)'\)", raw)
    codes += re.findall(r'"([a-zA-Z\-]+)"', raw)
    codes += re.findall(r'\b([a-zA-Z]{2,3}(?:-[A-Za-z0-9]+)?)\b', raw)

    return sorted(list(set(codes)))


def build_excel(keys, codes, save_path):
    wb = Workbook()
    ws = wb.active
    header = ["en", "en"] + codes
    ws.append(header)

    r = 2
    for k in keys:
        row = [k, k]
        for lang in codes:
            row.append(f'=GOOGLETRANSLATE($A{r}, "en", "{lang}")')
        ws.append(row)
        r += 1

    wb.save(save_path)


# --------------------------------------------------------
# FIXED JSON WRITER — ESCAPES QUOTES, FIXES MULTILINES
# --------------------------------------------------------
def write_clean_json(path, data):

    def clean(text):
        if text is None:
            return ""

        text = str(text)

        # Normalize newline types
        text = text.replace("\r\n", "\n")

        # Convert REAL Excel newlines → \n literal
        text = text.replace("\n", "\\n")

        # Escape double quotes
        text = text.replace('"', '\\"')

        # Trim accidental trailing spaces
        text = text.strip()

        return text

    with open(path, "w", encoding="utf-8") as f:
        f.write("{\n")
        items = list(data.items())

        for i, (k, v) in enumerate(items):
            k = clean(k)   # 🔥 fix newline inside keys
            v = clean(v)   # 🔥 fix newline inside values

            line = f'  "{k}": "{v}"'
            if i < len(items) - 1:
                line += ","

            f.write(line + "\n")

        f.write("}")


# --------------------------------------------------------
# EXCEL → JSON (FINAL FIXED VERSION)
# --------------------------------------------------------
def excel_to_json(folder_path, excel_file):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    language_codes = []
    for code in headers:
        if not code:
            continue
        if code == "en" and "en" not in language_codes:
            language_codes.append("en")
        elif code == "en":
            continue
        else:
            language_codes.append(code)

    keys = []
    values = {lang: [] for lang in language_codes}

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        if not key:
            continue

        keys.append(key)

        for lang in language_codes:
            col_index = headers.index(lang)
            val = row[col_index] if col_index < len(row) and row[col_index] else ""
            values[lang].append(val)

    # Create folder if needed
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Write each language JSON
    for lang in language_codes:
        data = {keys[i]: values[lang][i] for i in range(len(keys))}
        json_file = os.path.join(folder_path, f"{lang}.json")
        write_clean_json(json_file, data)

    return True


# --------------------------------------------------------
# GUI APPLICATION
# --------------------------------------------------------
def create_app():
    root = tk.Tk()
    root.title("Localization Tool")
    root.geometry("500x590")

    selected_file = tk.StringVar()
    log = tk.StringVar()

    def log_msg(msg):
        log.set(msg)

    def choose_dart():
        p = filedialog.askopenfilename(filetypes=[("Dart files", "*.dart")])
        if p:
            selected_file.set(p)
            log_msg("Dart file selected")

    def choose_json():
        p = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if p:
            selected_file.set(p)
            log_msg("JSON file selected")

    def convert_to_json():
        p = selected_file.get()
        if not p:
            messagebox.showerror("Error", "No file selected")
            return

        try:
            items = extract_dart_strings(p)
            save = filedialog.asksaveasfilename(defaultextension=".json")

            if save:
                raw = json.dumps({s: s for s in items}, ensure_ascii=False, indent=2)
                raw = raw.replace("\\\\n", "\\n")

                with open(save, "w", encoding="utf-8") as f:
                    f.write(raw)

                log_msg("JSON saved")
        except:
            messagebox.showerror("Error", "Conversion failed")

    def generate_excel():
        p = selected_file.get()
        if not p:
            messagebox.showerror("Error", "No file selected")
            return

        try:
            keys = extract_dart_strings(p) if p.endswith(".dart") else extract_json_keys(p)
            codes = ask_languages_popup(root)

            save = filedialog.asksaveasfilename(defaultextension=".xlsx")
            if save:
                build_excel(keys, codes, save)
                log_msg("Excel saved")
        except Exception as e:
            messagebox.showerror("Error", f"Failed:\n{e}")

    def excel_to_json_action():
        excel = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not excel:
            return

        folder = filedialog.askdirectory()
        if not folder:
            return

        try:
            excel_to_json(folder, excel)
            log_msg("JSONs exported successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed:\n{e}")

    ttk.Button(root, text="Select Dart File", command=choose_dart).pack(pady=10)
    ttk.Button(root, text="Select JSON File", command=choose_json).pack(pady=10)
    ttk.Button(root, text="Convert Dart → JSON", command=convert_to_json).pack(pady=10)
    ttk.Button(root, text="Generate translation.xlsx", command=generate_excel).pack(pady=10)
    ttk.Button(root, text="Convert Excel → JSONs", command=excel_to_json_action).pack(pady=10)

    tk.Label(root, textvariable=log, fg="green").pack(pady=20)

    root.mainloop()


create_app()
