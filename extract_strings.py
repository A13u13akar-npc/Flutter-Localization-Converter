import json
import re
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook


def choose():
    print("1. Process Dart file")
    print("2. Process JSON file")
    print("3. Convert Excel → JSONs")
    return input("Select: ").strip()


def select_file(ftype):
    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    if ftype == "dart":
        path = filedialog.askopenfilename(filetypes=[("Dart files", "*.dart")])
    elif ftype == "json":
        path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
    elif ftype == "excel":
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    else:
        path = ""
    root.destroy()
    return path


def select_save(ext):
    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    if ext == "json":
        path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
    else:
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    root.destroy()
    return path


def select_folder():
    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    path = filedialog.askdirectory()
    root.destroy()
    return path


def strip_comments(text):
    text = re.sub(r'/\*.*?\*/', '', text, flags=re.DOTALL)
    cleaned = []
    for line in text.splitlines():
        line = re.sub(r'//.*', '', line)
        if line.strip() != "":
            cleaned.append(line)
    return "\n".join(cleaned)


def normalize(s):
    s = s.replace("\\n", "\n")
    s = s.replace("\n", "\\n")
    return s


def extract_dart(text):
    text = strip_comments(text)
    pattern = r'"([^"\$]+)"\.tr'
    out = []
    for m in re.findall(pattern, text):
        out.append(normalize(m))
    return out


def extract_json_keys(path):
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    raw = strip_comments(raw)
    data = json.loads(raw)
    return list(data.keys())


def ask_langs():
    print("Paste the supported languages list below. Finish with an empty line:")
    lines = []
    while True:
        l = input()
        if l.strip() == "":
            break
        lines.append(l.strip())

    text = "\n".join(lines)
    text = re.sub(r'//.*', '', text)
    text = re.sub(r'/\*.*?\*/', '', text, flags=re.DOTALL)

    codes = []
    codes += re.findall(r"Locale\('([a-zA-Z\-]+)'\)", text)
    codes += re.findall(r'"([a-zA-Z\-]+)"', text)
    codes += re.findall(r'\b([a-zA-Z]{2,3}(?:-[A-Za-z0-9]+)?)\b', text)

    codes = [c for c in codes if len(c) >= 2]
    codes = sorted(list(set(codes)))

    print("Detected language codes:")
    for c in codes:
        print(c)

    return codes


def build_excel(keys, codes):
    wb = Workbook()
    ws = wb.active
    header = ["en", "en"] + codes
    ws.append(header)
    row_num = 2
    for k in keys:
        row = [k, k]
        for lang in codes:
            row.append('=GOOGLETRANSLATE($A' + str(row_num) + ', "en", "' + lang + '")')
        ws.append(row)
        row_num += 1
    return wb


def write_clean_json(path, data):
    def clean(text):
        if text is None:
            return ""
        text = str(text)
        text = text.replace("\r\n", "\n")
        text = text.replace("\n", "\\n")
        text = text.replace('"', '\\"')
        text = text.strip()
        return text

    with open(path, "w", encoding="utf-8") as f:
        f.write("{\n")
        items = list(data.items())
        for i, (k, v) in enumerate(items):
            k = clean(k)
            v = clean(v)
            line = f'  "{k}": "{v}"'
            if i < len(items) - 1:
                line += ","
            f.write(line + "\n")
        f.write("}")


def excel_to_json(folder_path, excel_file):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    language_codes = []
    for code in headers:
        if not code:
            continue
        if code == "en" and "en" not in language_codes:
            language_codes.append("en")
        elif code == "en":
            continue
        else:
            language_codes.append(code)

    keys = []
    values = {lang: [] for lang in language_codes}

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        if not key:
            continue
        keys.append(key)
        for lang in language_codes:
            col_index = headers.index(lang)
            val = row[col_index] if col_index < len(row) and row[col_index] else ""
            values[lang].append(val)

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    for lang in language_codes:
        data = {keys[i]: values[lang][i] for i in range(len(keys))}
        json_path = os.path.join(folder_path, f"{lang}.json")
        write_clean_json(json_path, data)

    return True


def handle_dart():
    print("Choose Dart file")
    path = select_file("dart")
    if not path:
        print("No file selected")
        return

    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    items = extract_dart(content)

    print("1. Print extracted strings")
    print("2. Convert to JSON")
    print("3. Generate translation.xlsx")
    choice = input("Select: ").strip()

    if choice == "1":
        for s in items:
            print(s)
        return

    if choice == "2":
        save = select_save("json")
        if not save:
            print("No destination selected")
            return
        data = {s: s for s in items}
        raw = json.dumps(data, ensure_ascii=False, indent=2)
        raw = raw.replace("\\\\n", "\\n")
        with open(save, "w", encoding="utf-8") as f:
            f.write(raw)
        print("Saved:", save)
        return

    if choice == "3":
        codes = ask_langs()
        wb = build_excel(items, codes)
        save = select_save("xlsx")
        if not save:
            print("No destination selected")
            return
        wb.save(save)
        print("Saved:", save)
        return


def handle_json():
    print("Choose JSON file")
    path = select_file("json")
    if not path:
        print("No file selected")
        return

    keys = extract_json_keys(path)

    print("1. Print all strings")
    print("2. Generate translation.xlsx")
    choice = input("Select: ").strip()

    if choice == "1":
        for k in keys:
            print(k)
        return

    if choice == "2":
        codes = ask_langs()
        wb = build_excel(keys, codes)
        save = select_save("xlsx")
        if not save:
            print("No destination selected")
            return
        wb.save(save)
        print("Saved:", save)
        return


def handle_excel_to_json():
    print("Choose Excel file (.xlsx)")
    excel = select_file("excel")
    if not excel:
        print("No file selected")
        return

    print("Choose output folder for JSON files")
    folder = select_folder()
    if not folder:
        print("No folder selected")
        return

    try:
        excel_to_json(folder, excel)
        print("JSONs exported successfully to:", folder)
    except Exception as e:
        print("Failed to convert Excel to JSONs:", e)


def main():
    c = choose()
    if c == "1":
        handle_dart()
    elif c == "2":
        handle_json()
    elif c == "3":
        handle_excel_to_json()
    else:
        print("Invalid option")


if __name__ == "__main__":
    main()
