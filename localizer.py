import json
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
import ttkbootstrap as tb
from ttkbootstrap.constants import *


# =========================================================
# UTILITIES
# =========================================================

def strip_comments(text):
    text = re.sub(r'/\*.*?\*/', "", text, flags=re.DOTALL)
    return "\n".join(
        [re.sub(r'//.*', "", l).strip() for l in text.splitlines() if l.strip()]
    )


def normalize(s):
    # Keep \n as literal in storage
    return s.replace("\\n", "\n").replace("\n", "\\n")


def extract_dart_strings(path):
    with open(path, "r", encoding="utf-8") as f:
        txt = strip_comments(f.read())
    return [normalize(x) for x in re.findall(r'"([^"\$]+)"\.tr', txt)]


def extract_json_keys(path):
    with open(path, "r", encoding="utf-8") as f:
        txt = strip_comments(f.read())
    return list(json.loads(txt).keys())


def ask_languages_popup(root):
    win = tb.Toplevel(root)
    win.title("Language Codes")
    win.attributes("-topmost", True)

    # Window size (responsive)
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    w, h = int(sw * 0.30), int(sh * 0.40)
    win.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    # --- Layout Frame ---
    outer = tb.Frame(win)
    outer.pack(expand=True, fill="both", padx=10, pady=10)

    result = tk.StringVar()

    def done(_=None):
        result.set(txt.get("1.0", "end"))
        win.destroy()

    # 🔥 OK BUTTON ON TOP (ALWAYS VISIBLE)
    tb.Button(
        outer,
        text="✔ Enter",
        bootstyle="light",
        command=done,
        width=18
    ).pack(pady=(0, 10))

    # TEXT FIELD BELOW
    txt = tk.Text(
        outer,
        bg="#111",
        fg="#fff",
        insertbackground="#fff",
        wrap="word"
    )
    txt.pack(expand=True, fill="both")

    root.wait_window(win)

    raw = re.sub(r'/\*.*?\*/', "", re.sub(r'//.*', "", result.get()), flags=re.DOTALL)

    codes = []
    codes += re.findall(r"Locale\('([a-zA-Z\-]+)'\)", raw)
    codes += re.findall(r'"([a-zA-Z\-]+)"', raw)
    codes += re.findall(r'\b([a-z]{2,3}(?:-[A-Za-z0-9]+)?)\b', raw)

    return sorted(set(codes))


# =========================================================
# BUILD EXCEL — EXACT FORMAT YOU ASKED FOR
# Col A: English source
# Col B: Editable English copy
# Col C..: GOOGLETRANSLATE($Arow, "en", "<lang>")
# =========================================================
def build_excel(keys, codes, save_path):
    wb = Workbook()
    ws = wb.active

    # Header: en (source), en (editable), then language codes
    ws.append(["en", "en"] + codes)

    row_num = 2
    for k in keys:
        # Col A: English source
        # Col B: Editable English copy
        row = [k, k]

        # Remaining columns: GOOGLETRANSLATE from column A
        for lang in codes:
            formula = f'=GOOGLETRANSLATE($A{row_num}, "en", "{lang}")'
            row.append(formula)

        ws.append(row)
        row_num += 1

    wb.save(save_path)


def write_clean_json(path, data):

    def clean(t):
        if t is None:
            return ""
        t = str(t).replace("\r\n", "\n").replace("\n", "\\n").replace('"', '\\"')
        return t.strip()

    with open(path, "w", encoding="utf-8") as f:
        f.write("{\n")
        items = list(data.items())
        for i, (k, v) in enumerate(items):
            comma = "," if i < len(items) - 1 else ""
            f.write(f'  "{clean(k)}": "{clean(v)}"{comma}\n')
        f.write("}")


def excel_to_json(folder, file):
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    headers = [(c.value or "").strip() for c in ws[1]]

    langs = []
    for h in headers:
        if h == "en" and "en" not in langs:
            langs.append("en")
        elif h not in ("", "en"):
            langs.append(h)

    keys = []
    values = {l: [] for l in langs}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        keys.append(row[0])

        for lang in langs:
            idx = headers.index(lang)
            values[lang].append(row[idx] if idx < len(row) and row[idx] is not None else "")

    os.makedirs(folder, exist_ok=True)

    for lang in langs:
        out = {keys[i]: values[lang][i] for i in range(len(keys))}
        write_clean_json(os.path.join(folder, f"{lang}.json"), out)


# =========================================================
# COMBINE JSON FOLDERS
# =========================================================
def combine_json_folders(old_folder, new_folder, output_folder):
    os.makedirs(output_folder, exist_ok=True)

    for filename in os.listdir(old_folder):
        if not filename.endswith(".json"):
            continue

        old_path = os.path.join(old_folder, filename)
        new_path = os.path.join(new_folder, filename)
        out_path = os.path.join(output_folder, filename)

        try:
            with open(old_path, "r", encoding="utf-8") as f:
                old_data = json.load(f)
        except Exception:
            old_data = {}

        try:
            with open(new_path, "r", encoding="utf-8") as f:
                new_data = json.load(f)
        except Exception:
            new_data = {}

        merged = dict(old_data)
        for k, v in new_data.items():
            if k not in merged:
                merged[k] = v

        write_clean_json(out_path, merged)


# =========================================================
# UI / APP
# =========================================================

def create_app():
    root = tb.Window(themename="darkly")
    root.title("Localization Tool")

    # ---------------------------------------------------
    # APP ICON (Windows + macOS)
    # ---------------------------------------------------

    # Windows icon
    try:
        root.iconbitmap("icon.ico")
    except:
        pass

    # macOS / PNG fallback
    try:
        icon_img = tk.PhotoImage(file="icon.png")
        root.iconphoto(True, icon_img)
    except:
        pass

    sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
    w, h = int(sw * 0.45), int(sh * 0.55)
    root.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    main = tb.Frame(root)
    main.pack(expand=True)

    selected_file = tk.StringVar()
    log_var = tk.StringVar()

    def big_btn(parent, text, command):
        return tb.Button(
            parent,
            text=text,
            command=command,
            bootstyle="secondary",
            width=30,
            padding=40,
        )

    def group(title):
        frame = tb.Labelframe(main, text=title, padding=12, bootstyle="warning")
        frame.pack(fill="x", pady=8)
        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)
        return frame

    # ---------- ACTIONS ----------

    def choose_dart():
        p = filedialog.askopenfilename(filetypes=[("Dart Files", "*.dart")])
        if p:
            selected_file.set(p)
            log_var.set("✔ Dart file selected")

    def choose_json():
        p = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json")])
        if p:
            selected_file.set(p)
            log_var.set("✔ JSON file selected")

    def convert_dart_json():
        p = selected_file.get()
        if not p:
            return messagebox.showerror("Error", "Select a Dart file first")
        try:
            items = extract_dart_strings(p)
            save = filedialog.asksaveasfilename(defaultextension=".json")
            if not save:
                return
            raw = json.dumps({s: s for s in items}, ensure_ascii=False, indent=2)
            raw = raw.replace("\\\\n", "\\n")
            with open(save, "w", encoding="utf-8") as f:
                f.write(raw)
            messagebox.showinfo("Done", "Dart → JSON complete")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def gen_excel():
        p = selected_file.get()
        if not p:
            return messagebox.showerror("Error", "Select a Dart/JSON file first")

        try:
            keys = extract_dart_strings(p) if p.endswith(".dart") else extract_json_keys(p)
            langs = ask_languages_popup(root)
            save = filedialog.asksaveasfilename(defaultextension=".xlsx")
            if not save:
                return
            build_excel(keys, langs, save)
            messagebox.showinfo("Done", "translation.xlsx created")
        except Exception as e:
            messagebox.showerror("Error", f"Excel generation failed:\n{e}")

    def excel_to_json_action():
        x = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not x:
            return
        folder = filedialog.askdirectory()
        if not folder:
            return
        try:
            excel_to_json(folder, x)
            messagebox.showinfo("Done", "JSONs exported")
        except Exception as e:
            messagebox.showerror("Error", f"Excel → JSONs failed:\n{e}")

    def combine_action():
        old = filedialog.askdirectory(title="Select OLD JSON folder")
        if not old:
            return
        new = filedialog.askdirectory(title="Select NEW JSON folder")
        if not new:
            return
        out = filedialog.askdirectory(title="Select OUTPUT folder")
        if not out:
            return
        try:
            combine_json_folders(old, new, out)
            messagebox.showinfo("Done", "Merged JSONs saved")
        except Exception as e:
            messagebox.showerror("Error", f"Combine failed:\n{e}")

    # ---------- LAYOUT ----------

    grp1 = group("🔤INPUT FILES")
    big_btn(grp1, "📥 Select Dart File", choose_dart).grid(row=0, column=0, padx=8, pady=6)
    big_btn(grp1, "📥 Select JSON File", choose_json).grid(row=0, column=1, padx=8, pady=6)

    grp2 = group("🌍TRANSLATION")
    big_btn(grp2, "📊 Generate translation.xlsx", gen_excel).grid(row=0, column=0, padx=8, pady=6)
    big_btn(grp2, "🔀 Combine JSON Folders", combine_action).grid(row=0, column=1, padx=8, pady=6)

    grp3 = group("🔁CONVERTERS")
    big_btn(grp3, "🔁 Dart → JSON", convert_dart_json).grid(row=0, column=0, padx=8, pady=6)
    big_btn(grp3, "📂 Excel → JSONs", excel_to_json_action).grid(row=0, column=1, padx=8, pady=6)

    tb.Label(main, textvariable=log_var, font=("Segoe UI", 12), bootstyle="success").pack(pady=10)

    root.mainloop()


create_app()
